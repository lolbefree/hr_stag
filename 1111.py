import openpyxl
from datetime import datetime
from dateutil import relativedelta

wb = openpyxl.load_workbook('Стаж роботи працівників на Укрлада.xlsx')
sheets = wb.sheetnames
sheet = wb["Стаж"]

list1 = []

for row in sheet.rows:
    list1.append(row[2].value)

_month_ = []
_years_ = []
res_list = []
over_list = []


all_rows = []

for row in sheet:
    current_row = []
    for cell in row:
        current_row.append(cell.value)
    all_rows.append(current_row)
name = []
for i in all_rows:
    if "ПІБ" not in i and "Дата прийняття" not in i:
        name.append(i[1])
income_date = []

for i in all_rows:
    try:
        if "Дата прийняття" not in i and i != None:
            # print(type(i[2]))
           income_date.append(i[2].strftime("%d.%m.%Y"))
    except AttributeError:
        continue
for _date_ in list1:
    if isinstance(_date_, datetime):
        today = datetime.now()
        r = relativedelta.relativedelta(today, _date_)
        _years_.append(r.years)
        _month_.append(r.months)

for year, month in zip(_years_, _month_):
    res_list.append("{} р.{} міс.".format(year, month))

less_year = []
from_one_year_to_three = []
from_three_year_to_ten = []
after_ten_years = []

cnt_more_then_10 = 0
cnt_from_3_to_10 = 0
cnt_from_1_to_3 = 0
cnt_before_1 = 0

try:
    wb.remove(wb['до 1 року'])
    wb.remove(wb['від 1 до 3 років'])
    wb.remove(wb['Більше 10 років'])
    wb.remove(wb['від 3 до 10 років'])
except:
    pass
w4 = wb.create_sheet("до 1 року", 2)
w3 = wb.create_sheet("від 1 до 3 років", 3)
w2 = wb.create_sheet("від 3 до 10 років", 4)

w1 = wb.create_sheet("Більше 10 років", 5)
w1.cell(row=1, column=1).value = "Піб"
w1.cell(row=1, column=2).value = "Дата прийняття"
w1.cell(row=1, column=3).value = "Стаж роботи на Автосоюз"
w1.cell(row=1, column=3).value = "Стаж роботи на Автосоюз"
w1.cell(row=1, column=4).value = "Оплата ДМС працівником"
w1.column_dimensions['A'].width = 40
w1.column_dimensions['B'].width = 15
w1.column_dimensions['C'].width = 24
w1.column_dimensions['D'].width = 24

w2.cell(row=1, column=1).value = "Піб"
w2.cell(row=1, column=2).value = "Дата прийняття"
w2.cell(row=1, column=3).value = "Стаж роботи на Автосоюз"
w2.cell(row=1, column=3).value = "Стаж роботи на Автосоюз"
w2.cell(row=1, column=4).value = "Оплата ДМС працівником"
w2.column_dimensions['A'].width = 40
w2.column_dimensions['B'].width = 15
w2.column_dimensions['C'].width = 24
w2.column_dimensions['D'].width = 24
w3.cell(row=1, column=1).value = "Піб"
w3.cell(row=1, column=2).value = "Дата прийняття"
w3.cell(row=1, column=3).value = "Стаж роботи на Автосоюз"
w3.cell(row=1, column=3).value = "Стаж роботи на Автосоюз"
w3.cell(row=1, column=4).value = "Оплата ДМС працівником"
w3.column_dimensions['A'].width = 40
w3.column_dimensions['B'].width = 15
w3.column_dimensions['C'].width = 24
w3.column_dimensions['D'].width = 24
w4.cell(row=1, column=1).value = "Піб"
w4.cell(row=1, column=2).value = "Дата прийняття"
w4.cell(row=1, column=3).value = "Стаж роботи на Автосоюз"
w4.cell(row=1, column=3).value = "Стаж роботи на Автосоюз"
w4.cell(row=1, column=4).value = "Оплата ДМС працівником"
row1 = 2
w4.column_dimensions['A'].width = 40
w4.column_dimensions['B'].width = 15
w4.column_dimensions['C'].width = 24
w4.column_dimensions['D'].width = 24

for i, j, k in zip(name, income_date, res_list):

    if int(res_list[cnt_more_then_10][:2].replace(" ", "")) >= 10:
        after_ten_years.append(i)
        after_ten_years.append(j)
        after_ten_years.append(k)

    cnt_more_then_10 += 1

    if 3 <= int(res_list[cnt_from_3_to_10][:2].replace(" ", "")) < 10:
        from_three_year_to_ten.append(i)
        from_three_year_to_ten.append(j)
        from_three_year_to_ten.append(k)
    cnt_from_3_to_10 += 1

    if 1 <= int(res_list[cnt_from_1_to_3][:2].replace(" ", "")) < 3:
        from_one_year_to_three.append(i)
        from_one_year_to_three.append(j)
        from_one_year_to_three.append(k)
    cnt_from_1_to_3 += 1
    if int(res_list[cnt_before_1][:2].replace(" ", "")) < 1:
        less_year.append(i)
        less_year.append(j)
        less_year.append(k)
    cnt_before_1 += 1

# print(less_year)
for item in after_ten_years[::3]:
    w1.cell(row=row1, column=1).value = str(item)
    row1 += 1
row1 = 2
tmp = 1     # не забываем прыгать по рядам
try:
    for item in after_ten_years:
        w1.cell(row=row1, column=2).value = (str(after_ten_years[tmp]))
        tmp += 3
        row1 += 1
        # w1.cell(row=row1, column=3).value = str(after_ten_years[stage])
except:
    pass
row1 = 2
tmp = 2     # не забываем прыгать по рядам
try:
    for item in after_ten_years:
        w1.cell(row=row1, column=3).value = (str(after_ten_years[tmp]))
        tmp += 3
        row1 += 1
        # w1.cell(row=row1, column=3).value = str(after_ten_years[stage])
except:
    pass
row1 = 2
try:
    for item in after_ten_years[::3]:
        w1.cell(row=row1, column=4).value = "30 %"
        # w1.cell(row=row1, column=3).value = str(after_ten_years[stage])
        row1 += 1
except:
    pass




# # from 3 to 10


row1 = 2
for item in from_three_year_to_ten[::3]:
    w2.cell(row=row1, column=1).value = str(item)
    row1 += 1
row1 = 2
tmp = 1     # не забываем прыгать по рядам
try:
    for item in from_three_year_to_ten:
        w2.cell(row=row1, column=2).value = (str(from_three_year_to_ten[tmp]))
        tmp += 3
        row1 += 1
        # w1.cell(row=row1, column=3).value = str(after_ten_years[stage])
except:
    pass
row1 = 2
tmp = 2     # не забываем прыгать по рядам
try:
    for item in from_three_year_to_ten:
        w2.cell(row=row1, column=3).value = (str(from_three_year_to_ten[tmp]))
        tmp += 3
        row1 += 1
        # w1.cell(row=row1, column=3).value = str(after_ten_years[stage])
except:
    pass
row1 = 2
try:
    for item in from_three_year_to_ten[::3]:
        w2.cell(row=row1, column=4).value = "50 %"
        # w1.cell(row=row1, column=3).value = str(after_ten_years[stage])
        row1 += 1
except:
    pass

# from 1 to 3

row1 = 2
for item in from_one_year_to_three[::3]:
    print(item)
    w3.cell(row=row1, column=1).value = str(item)
    row1 += 1
row1 = 2
tmp = 1     # не забываем прыгать по рядам
try:
    for item in from_one_year_to_three:
        w3.cell(row=row1, column=2).value = (str(from_one_year_to_three[tmp]))
        tmp += 3
        row1 += 1
        # w1.cell(row=row1, column=3).value = str(after_ten_years[stage])
except:
    pass
row1 = 2
tmp = 2     # не забываем прыгать по рядам
try:
    for item in from_one_year_to_three:
        w3.cell(row=row1, column=3).value = (str(from_one_year_to_three[tmp]))
        tmp += 3
        row1 += 1
        # w1.cell(row=row1, column=3).value = str(after_ten_years[stage])
except:
    pass
row1 = 2
try:
    for item in from_one_year_to_three[::3]:
        w3.cell(row=row1, column=4).value = "70 %"
        # w1.cell(row=row1, column=3).value = str(after_ten_years[stage])
        row1 += 1
except:
    pass

# before 1


row1 = 2
for item in less_year[::3]:
    # print(item)
    w4.cell(row=row1, column=1).value = str(item)
    row1 += 1
row1 = 2
tmp = 1     # не забываем прыгать по рядам
try:
    for item in less_year:
        w4.cell(row=row1, column=2).value = (str(less_year[tmp]))
        tmp += 3
        row1 += 1
        # w1.cell(row=row1, column=3).value = str(after_ten_years[stage])
except:
    pass
row1 = 2
tmp = 2     # не забываем прыгать по рядам
try:
    for item in less_year:
        w4.cell(row=row1, column=3).value = (str(less_year[tmp]))
        tmp += 3
        row1 += 1
        # w1.cell(row=row1, column=3).value = str(after_ten_years[stage])
except:
    pass
row1 = 2
try:
    for item in less_year[::3]:
        w4.cell(row=row1, column=4).value = "100 %"
        # w1.cell(row=row1, column=3).value = str(after_ten_years[stage])
        row1 += 1
except:
    pass

# print(less_year)

print(len(less_year)/3, len(from_one_year_to_three)/3, len(from_three_year_to_ten)/3, len(after_ten_years)/3)
# print(i, j, k)
del _month_
del _years_
wb.save("Стаж роботи працівників на Укрлада.xlsx")
